"""
XLSX 解析器 — 基于 openpyxl。

将 Excel 表格内容解析为可检索的 Document 列表。
支持按 sheet、按行、按指定列等多种模式。
"""

import hashlib
from typing import List, Optional

from langchain_core.documents import Document
from tqdm import tqdm

from kb.offline.scripts.file_processors.base_parser import BaseParser


class XLSXParser(BaseParser):
    """Excel (.xlsx / .xls) 格式解析器。"""

    def supported_extensions(self) -> List[str]:
        return [".xlsx", ".xls"]

    def parse(
        self,
        file_path: str,
        mode: str = "sheet",
        sheet_name: Optional[str] = None,
        header_row: Optional[int] = 0,
        **kwargs,
    ) -> List[Document]:
        """
        解析 Excel 文件。

        Args:
            file_path: 文件路径
            mode: 解析模式
                - "sheet": 每个 sheet 一个 Document
                - "row":   每行一个 Document（适合数据类表格）
                - "chunk": 按固定行数分块
            sheet_name: 指定 sheet（None 表示全部）
            header_row: 表头行号（0-index），None 表示无表头

        Returns:
            List[Document]: 解析结果
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "请安装 openpyxl: pip install openpyxl"
            )

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        raw_docs: List[Document] = []

        sheet_names = [sheet_name] if sheet_name else wb.sheetnames

        for name in tqdm(sheet_names, desc="解析 XLSX sheet"):
            ws = wb[name]
            if mode == "sheet":
                doc = self._parse_as_sheet(ws, name, file_path)
                if doc:
                    raw_docs.append(doc)
            elif mode == "row":
                raw_docs.extend(
                    self._parse_as_rows(ws, name, file_path, header_row)
                )
            elif mode == "chunk":
                raw_docs.extend(
                    self._parse_as_chunks(ws, name, file_path, header_row)
                )

        wb.close()
        print(f"[XLSXParser] 解析完成: {file_path} → {len(raw_docs)} 个文档块")
        return raw_docs

    def _parse_as_sheet(
        self, ws, sheet_name: str, file_path: str
    ) -> Optional[Document]:
        """将整个 sheet 转为 Markdown 格式的表格文本。"""
        rows_data = []
        for row in ws.iter_rows(values_only=True):
            # 过滤全部为 None 的行
            if any(cell is not None for cell in row):
                row_text = " | ".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                rows_data.append(row_text)

        if not rows_data:
            return None

        content = "\n".join(rows_data)
        unique_id = hashlib.md5(
            f"{file_path}:{sheet_name}".encode("utf-8")
        ).hexdigest()
        metadata = {
            "unique_id": unique_id,
            "source": file_path,
            "format": "xlsx",
            "sheet": sheet_name,
        }
        return Document(page_content=content, metadata=metadata)

    def _parse_as_rows(
        self,
        ws,
        sheet_name: str,
        file_path: str,
        header_row: Optional[int] = 0,
    ) -> List[Document]:
        """每行一个 Document，首行作为表头。"""
        headers = []
        rows_iter = ws.iter_rows(values_only=True)

        # 读取表头
        if header_row is not None:
            try:
                header_cells = next(rows_iter)
                headers = [
                    str(c) if c is not None else f"col_{i}"
                    for i, c in enumerate(header_cells)
                ]
            except StopIteration:
                return []

        # 每行一个 Document
        docs = []
        for row_idx, row in enumerate(rows_iter):
            if all(cell is None for cell in row):
                continue
            if headers:
                pairs = []
                for h, c in zip(headers, row):
                    val = str(c) if c is not None else ""
                    pairs.append(f"{h}: {val}")
                content = "\n".join(pairs)
            else:
                content = " | ".join(
                    str(c) if c is not None else "" for c in row
                )

            unique_id = hashlib.md5(
                f"{file_path}:{sheet_name}:row{row_idx}".encode("utf-8")
            ).hexdigest()
            metadata = {
                "unique_id": unique_id,
                "source": file_path,
                "format": "xlsx",
                "sheet": sheet_name,
                "row_index": row_idx,
            }
            docs.append(Document(page_content=content, metadata=metadata))

        return docs

    def _parse_as_chunks(
        self,
        ws,
        sheet_name: str,
        file_path: str,
        header_row: Optional[int] = 0,
        chunk_size: int = 20,
    ) -> List[Document]:
        """按行数分块，每块一个 Document。"""
        headers = []
        rows_iter = ws.iter_rows(values_only=True)

        if header_row is not None:
            try:
                header_cells = next(rows_iter)
                headers = [
                    str(c) if c is not None else f"col_{i}"
                    for i, c in enumerate(header_cells)
                ]
            except StopIteration:
                return []

        docs = []
        current_chunk = []
        chunk_idx = 0

        for row_idx, row in enumerate(rows_iter):
            if all(cell is None for cell in row):
                continue
            if headers:
                pairs = []
                for h, c in zip(headers, row):
                    val = str(c) if c is not None else ""
                    pairs.append(f"{h}: {val}")
                current_chunk.append(" | ".join(pairs))
            else:
                current_chunk.append(
                    " | ".join(str(c) if c is not None else "" for c in row)
                )

            if len(current_chunk) >= chunk_size:
                content = "\n".join(current_chunk)
                unique_id = hashlib.md5(
                    f"{file_path}:{sheet_name}:chunk{chunk_idx}".encode("utf-8")
                ).hexdigest()
                metadata = {
                    "unique_id": unique_id,
                    "source": file_path,
                    "format": "xlsx",
                    "sheet": sheet_name,
                    "chunk_index": chunk_idx,
                }
                docs.append(Document(page_content=content, metadata=metadata))
                current_chunk = []
                chunk_idx += 1

        # 剩余行
        if current_chunk:
            content = "\n".join(current_chunk)
            unique_id = hashlib.md5(
                f"{file_path}:{sheet_name}:chunk{chunk_idx}".encode("utf-8")
            ).hexdigest()
            metadata = {
                "unique_id": unique_id,
                "source": file_path,
                "format": "xlsx",
                "sheet": sheet_name,
                "chunk_index": chunk_idx,
            }
            docs.append(Document(page_content=content, metadata=metadata))

        return docs


# ---- 便利函数 ----
def load_xlsx(file_path: str, **kwargs) -> List[Document]:
    """便利函数，可直接调用。"""
    return XLSXParser().parse(file_path, **kwargs)
